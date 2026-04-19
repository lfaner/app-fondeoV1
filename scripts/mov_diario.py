"""
Descarga el reporte 'Registro de ingresos/egresos' desde la UI de Aunesa
y muestra los resultados por consola. Sin carga a PostgreSQL.
"""
import os
import sys
import re
import json
import tempfile
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from dotenv import load_dotenv

BASE_DIR = Path(__file__).parent.parent
load_dotenv(BASE_DIR / '.env')

USERNAME = os.environ.get('AUNESA_USERNAME', '')
PASSWORD = os.environ.get('AUNESA_PASSWORD', '')
HOST_URL = os.environ.get('AUNESA_HOST', 'https://becerra.aunesa.com/Irmo') + '/'
MEP = float(os.environ.get('USD_MEP', 1168))
CCL = float(os.environ.get('USD_CCL', 1180))

config_path = BASE_DIR / 'fechas_config.json'
try:
    fechas = json.loads(config_path.read_text(encoding='utf-8'))
    fecha_desde = fechas.get('fecha_desde', '')
    fecha_hasta = fechas.get('fecha_hasta', '')
except FileNotFoundError:
    print('ERROR: fechas_config.json no encontrado.')
    sys.exit(1)

try:
    fecha_desde_ui = datetime.strptime(fecha_desde, '%Y-%m-%d').strftime('%d/%m/%Y')
    fecha_hasta_ui = datetime.strptime(fecha_hasta, '%Y-%m-%d').strftime('%d/%m/%Y')
except ValueError as e:
    print(f'ERROR en formato de fechas: {e}')
    sys.exit(1)

print(f'Período: {fecha_desde_ui} — {fecha_hasta_ui}')


def descargar_excel(headless=True):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        print('Abriendo Aunesa...')
        page.goto(HOST_URL)

        print('Login...')
        page.fill('.v-textfield-username', USERNAME)
        page.fill('.v-textfield-pass', PASSWORD)
        page.get_by_text('Ingresar / Log In').click()
        page.wait_for_selector('text=Consultas', timeout=20000)

        print('Navegando a Informes de auditoría...')
        buscador = page.locator('.v-filterselect-input').nth(1)
        buscador.click()
        buscador.fill('')
        buscador.type('Informes de auditor', delay=80)

        opened = False
        for option_text in (
            'Informes de auditoría (AN/AP) :: Gestión bursátil',
            'Informes de auditoria (AN/AP) :: Gestion bursatil',
            'Informes de auditorÃ­a (AN/AP) :: GestiÃ³n bursÃ¡til',
        ):
            try:
                page.get_by_text(option_text).first.click(timeout=3000)
                opened = True
                break
            except Exception:
                pass

        if not opened:
            page.keyboard.press('ArrowDown')
            page.keyboard.press('Enter')

        page.wait_for_timeout(3000)
        page.wait_for_selector('.v-datefield-textfield')

        fecha_inputs = page.locator('.v-datefield-textfield')
        fecha_inputs.nth(0).click(click_count=3)
        fecha_inputs.nth(0).type(fecha_desde_ui, delay=40)
        page.keyboard.press('Tab')

        fecha_inputs.nth(1).click(click_count=3)
        fecha_inputs.nth(1).type(fecha_hasta_ui, delay=40)
        page.keyboard.press('Tab')

        concertacion = page.locator("input[type='checkbox']").first
        if not concertacion.is_checked():
            concertacion.check()

        page.wait_for_timeout(1000)

        print('Descargando Registro de ingresos/egresos...')
        with page.expect_download(timeout=180000) as dl:
            page.get_by_role('button', name='Registro de ingresos/egresos').click()
        download = dl.value

        tmp = tempfile.mktemp(suffix='.xlsx')
        download.save_as(tmp)
        print(f'Excel guardado temporalmente en: {tmp}')
        browser.close()
        return tmp


def parse_cantidad(valor):
    """Retorna el valor numérico tal como viene (negativo=débito, positivo=crédito)."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    txt = re.sub(r'\s*[AaDd]$', '', str(valor).strip())
    txt = txt.replace('.', '').replace(',', '.')
    try:
        return float(txt)
    except ValueError:
        return 0.0


def detectar_headers(ws):
    for r in range(1, 4):
        vals = [str(ws.cell(r, c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]
        if 'cuenta' in vals and 'unidad' in vals:
            return r, vals
    raise ValueError('No se encontraron headers esperados en el Excel.')


def procesar_excel(ruta):
    wb = load_workbook(ruta)
    ws = wb[wb.sheetnames[0]]

    header_row, headers = detectar_headers(ws)
    col = {name: idx + 1 for idx, name in enumerate(headers)}

    col_comprobante  = col.get('comprobante')
    col_liquidacion  = col.get('liquidación') or col.get('liquidacion')
    col_concertacion = col.get('concertación') or col.get('concertacion')
    col_cuenta       = col.get('cuenta')
    col_denominacion = col.get('denominación') or col.get('denominacion')
    col_unidad       = col.get('unidad')
    col_informacion  = col.get('información') or col.get('informacion')
    col_referencia   = col.get('referencia')
    col_cantidad     = col.get('cantidad')

    filas = []
    for r in range(header_row + 1, ws.max_row + 1):
        cuenta = ws.cell(r, col_cuenta).value if col_cuenta else None
        if not cuenta:
            continue

        liq_val = ws.cell(r, col_liquidacion).value if col_liquidacion else None
        if isinstance(liq_val, datetime):
            fecha = liq_val.strftime('%d/%m/%Y %H:%M:%S')
        else:
            fecha = str(liq_val or '').strip()

        unidad   = str(ws.cell(r, col_unidad).value or '').strip().upper() if col_unidad else ''
        cantidad = parse_cantidad(ws.cell(r, col_cantidad).value if col_cantidad else None)

        if unidad == 'USD':
            monto_ars = cantidad * MEP      # signo se preserva
            monto_usd = cantidad
        elif unidad == 'USDC':
            monto_ars = cantidad * CCL
            monto_usd = cantidad * CCL / MEP
        else:  # ARS
            monto_ars = cantidad
            monto_usd = cantidad / MEP

        filas.append({
            'comprobante':  str(ws.cell(r, col_comprobante).value or '').strip() if col_comprobante else '',
            'liquidacion':  fecha,
            'cuenta':       str(cuenta).strip(),
            'denominacion': str(ws.cell(r, col_denominacion).value or '').strip() if col_denominacion else '',
            'unidad':       unidad,
            'informacion':  str(ws.cell(r, col_informacion).value or '').strip() if col_informacion else '',
            'referencia':   str(ws.cell(r, col_referencia).value or '').strip() if col_referencia else '',
            'cantidad':     cantidad,
            'monto_ars':    monto_ars,
            'monto_usd':    monto_usd,
        })

    wb.close()
    return filas


# ── Main ──────────────────────────────────────────────────────────────────────
ruta_excel = descargar_excel(headless=False)

filas = procesar_excel(ruta_excel)
print(f'\nMEP: {MEP:,.2f}  |  CCL: {CCL:,.2f}')
print(f'Total filas: {len(filas)}\n')
print(f'{"Comprobante":<18} {"Liquidacion":<22} {"Cuenta":<10} {"Unidad":<6} {"Cantidad":>14} {"Monto ARS":>16} {"Monto USD":>14}  Referencia')
print('-' * 120)
for f in filas:
    print(
        f'{f["comprobante"]:<18} {f["liquidacion"]:<22} {f["cuenta"]:<10} '
        f'{f["unidad"]:<6} {f["cantidad"]:>14,.2f} {f["monto_ars"]:>16,.2f} {f["monto_usd"]:>14,.2f}  {f["referencia"]}'
    )
